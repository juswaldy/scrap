from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import math
import re
import sys
import time
import urllib.request
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np

try:
    from skyfield import almanac
    from skyfield.api import load
except ImportError:  # pragma: no cover - runtime dependency check
    almanac = None
    load = None

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - runtime dependency check
    Workbook = None
    get_column_letter = None

YEAR_START = 2001
YEAR_END = 2100

MONTHS = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}

WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

SOURCE_URLS = {
    "astropixels_moonperap2001.html": "https://www.astropixels.com/ephemeris/moon/moonperap2001.html",
    "astropixels_phases2001gmt.html": "https://www.astropixels.com/ephemeris/moon/phases2001gmt.html",
    "nasa_solar_eclipses_2001_2100.html": "https://eclipse.gsfc.nasa.gov/SEcat5/SE2001-2100.html",
    "nasa_lunar_eclipses_2001_2100.html": "https://eclipse.gsfc.nasa.gov/LEcat5/LE2001-2100.html",
}

JPL_DE421_URL = "https://ssd.jpl.nasa.gov/ftp/eph/planets/bsp/de440s.bsp"
JPL_DE421_FILENAME = "jpl_de440s.bsp"

# Major meteor showers — approximate peak month/day, peak solar longitude
# (J2000 ecliptic, degrees), and typical ZHR from IMO observations.  The peak
# solar longitude is used with skyfield to compute the annual peak for every
# year in the YEAR_START–YEAR_END range.
# Fields: (name, approx_month, approx_day, peak_solar_longitude_deg, typical_zhr)
METEOR_SHOWERS: List[Tuple[str, int, int, float, int]] = [
    ("Quadrantids",   1,  3, 283.16, 120),
    ("Lyrids",        4, 22,  32.32,  18),
    ("EAquarids",     5,  5,  45.5,   50),
    ("DAquarids",     7, 29, 125.0,   20),
    ("Perseids",      8, 12, 140.0,   80),
    ("Draconids",    10,  8, 195.4,    5),
    ("Orionids",     10, 21, 208.0,   25),
    ("Taurids",      11,  5, 220.0,   10),
    ("Leonids",      11, 18, 235.27,  16),
    ("Geminids",     12, 14, 262.2,  120),
]


class AstroDataError(RuntimeError):
    pass


def fetch_and_save_sources(download_dir: Path) -> Dict[str, str]:
    download_dir.mkdir(parents=True, exist_ok=True)
    out: Dict[str, str] = {}
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    total = len(SOURCE_URLS)
    for i, (filename, url) in enumerate(SOURCE_URLS.items(), 1):
        print(f"  [{i}/{total}] Downloading {filename} ...", end=" ", flush=True)
        target = download_dir / filename
        request = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(request) as response:
            charset = response.headers.get_content_charset() or "utf-8"
            raw_bytes = response.read()
        text = raw_bytes.decode(charset, errors="replace")
        target.write_text(text, encoding="utf-8")
        out[filename] = text
        print(f"done ({len(raw_bytes):,} bytes)")

    return out


def fetch_binary_with_headers(url: str, target: Path) -> None:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36",
        "Accept": "*/*",
    }
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request) as response:
        raw_bytes = response.read()
    target.write_bytes(raw_bytes)


def ensure_jpl_ephemeris(download_dir: Path, offline: bool) -> Path:
    ephemeris_path = download_dir / JPL_DE421_FILENAME
    if ephemeris_path.exists():
        return ephemeris_path

    if offline:
        raise AstroDataError(
            f"Offline mode requested, but {JPL_DE421_FILENAME} is missing in {download_dir}. "
            "Run once without --offline to download it."
        )

    fetch_binary_with_headers(JPL_DE421_URL, ephemeris_path)
    return ephemeris_path


def load_sources_offline(download_dir: Path) -> Dict[str, str]:
    out: Dict[str, str] = {}
    missing: List[str] = []

    for filename in SOURCE_URLS:
        path = download_dir / filename
        if not path.exists():
            missing.append(filename)
            continue
        out[filename] = path.read_text(encoding="utf-8")

    if missing:
        missing_csv = ", ".join(missing)
        raise AstroDataError(
            "Offline mode requested, but required download files are missing: "
            f"{missing_csv}. Run without --offline first to populate downloads."
        )

    return out


def strip_html_to_text(html_text: str) -> str:
    text = re.sub(r"<script[\s\S]*?</script>", " ", html_text, flags=re.IGNORECASE)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = text.replace("\xa0", " ")
    return text


def parse_nasa_solar_types(nasa_solar_html: str) -> Dict[Tuple[int, int, int], str]:
    text = strip_html_to_text(nasa_solar_html)
    pattern = re.compile(
        r"\b(20\d{2}|2100)\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{2})\s+"
        r"(\d{2}:\d{2}:\d{2})\s+\d+\s+\d+\s+\d+\s+([A-Z][A-Za-z0-9+\-]*)\b"
    )
    solar_map: Dict[Tuple[int, int, int], str] = {}

    for m in pattern.finditer(text):
        year = int(m.group(1))
        month = MONTHS[m.group(2)]
        day = int(m.group(3))
        token = m.group(5)
        t = token[0]
        if t == "T":
            code = "T"
        elif t == "A":
            code = "A"
        elif t == "P":
            code = "P"
        elif t == "H":
            code = "A"
        else:
            continue
        solar_map[(year, month, day)] = code

    return solar_map


def parse_nasa_lunar_types(nasa_lunar_html: str) -> Dict[Tuple[int, int, int], str]:
    text = strip_html_to_text(nasa_lunar_html)
    pattern = re.compile(
        r"\b(20\d{2}|2100)\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{2})\s+"
        r"(\d{2}:\d{2}:\d{2})\s+\d+\s+\d+\s+\d+\s+([A-Z][A-Za-z0-9+\-]*)\b"
    )
    lunar_map: Dict[Tuple[int, int, int], str] = {}

    for m in pattern.finditer(text):
        year = int(m.group(1))
        month = MONTHS[m.group(2)]
        day = int(m.group(3))
        token = m.group(5)
        t = token[0]
        if t == "T":
            code = "t"
        elif t == "P":
            code = "p"
        elif t == "N":
            code = "n"
        else:
            continue
        lunar_map[(year, month, day)] = code

    return lunar_map


def date_label(year: int, month: int, day: int) -> str:
    d = dt.date(year, month, day)
    weekday = WEEKDAYS[d.weekday()]
    return f"{weekday}, {d:%Y-%m-%d}"


def parse_phase_lines(phases_html: str) -> Dict[int, Dict[str, List[dict]]]:
    phase_keys = ["new", "first", "full", "last"]
    by_year: Dict[int, Dict[str, List[dict]]] = defaultdict(lambda: {k: [] for k in phase_keys})

    event_pattern = re.compile(
        r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})\s+(\d{2}:\d{2})(?:\s+([A-Za-z]))?(?=\s{2,}|$)"
    )

    pre_blocks = re.findall(r"<pre>([\s\S]*?)</pre>", phases_html, flags=re.IGNORECASE)

    for block in pre_blocks:
        lines = html.unescape(block).replace("\xa0", " ").splitlines()
        current_year: Optional[int] = None

        for raw_line in lines:
            line = raw_line.rstrip()
            year_match = re.match(r"^\s*(20\d{2}|2100)\b", line)
            if year_match:
                current_year = int(year_match.group(1))

            if not current_year or current_year < YEAR_START or current_year > YEAR_END:
                continue

            for match in event_pattern.finditer(line):
                month = MONTHS[match.group(1)]
                day = int(match.group(2))
                time_utc = match.group(3)
                raw_code = match.group(4) or ""
                start = match.start()

                if start < 19:
                    phase = "new"
                elif start < 37:
                    phase = "first"
                elif start < 55:
                    phase = "full"
                else:
                    phase = "last"

                by_year[current_year][phase].append(
                    {
                        "date": (current_year, month, day),
                        "time": time_utc,
                        "raw_code": raw_code,
                    }
                )

    return by_year


def normalize_phase_eclipse(
    phase: str,
    date_tuple: Tuple[int, int, int],
    raw_code: str,
    solar_map: Dict[Tuple[int, int, int], str],
    lunar_map: Dict[Tuple[int, int, int], str],
) -> str:
    if phase == "new":
        if date_tuple in solar_map:
            return solar_map[date_tuple]
        c = raw_code[:1].upper()
        if c == "H":
            return "A"
        return c if c in {"T", "P", "A"} else ""

    if phase == "full":
        if date_tuple in lunar_map:
            return lunar_map[date_tuple]
        c = raw_code[:1].upper()
        if c == "T":
            return "t"
        if c == "P":
            return "p"
        if c == "N":
            return "n"
        return ""

    return ""


def parse_perigee_apogee_rows(moonperap_html: str) -> List[dict]:
    text = strip_html_to_text(moonperap_html)
    lines = text.splitlines()

    rows: List[dict] = []

    def parse_event(tokens: List[str], idx: int) -> Tuple[Optional[dict], int]:
        if idx + 3 >= len(tokens):
            return None, idx
        mon = tokens[idx]
        if mon not in MONTHS:
            return None, idx

        day = tokens[idx + 1]
        t = tokens[idx + 2]
        dist = tokens[idx + 3]

        if not re.fullmatch(r"\d{1,2}", day):
            return None, idx
        if not re.fullmatch(r"\d{2}:\d{2}", t):
            return None, idx
        if not re.fullmatch(r"\d{6}", dist):
            return None, idx

        idx += 4

        source_flag = ""
        if idx < len(tokens) and tokens[idx] in {"m", "M"}:
            source_flag = tokens[idx]
            idx += 1

        if idx < len(tokens) and re.fullmatch(r"\d+\.\d+", tokens[idx]):
            idx += 1

        event = {
            "month": MONTHS[mon],
            "day": int(day),
            "time": t,
            "distance": int(dist),
            "source_flag": source_flag,
        }
        return event, idx

    current_year: Optional[int] = None

    for raw_line in lines:
        line = raw_line.rstrip()
        if not re.search(r"\d{2}:\d{2}", line):
            continue

        tokens = line.split()
        if not tokens:
            continue

        idx = 0
        if re.fullmatch(r"\d{4}", tokens[0]):
            current_year = int(tokens[0])
            idx = 1

        if current_year is None or current_year < YEAR_START or current_year > YEAR_END:
            continue

        first_event, idx2 = parse_event(tokens, idx)
        if first_event is None:
            continue

        second_event, _ = parse_event(tokens, idx2)

        perigee_event: Optional[dict] = None
        apogee_event: Optional[dict] = None

        if first_event["distance"] >= 390000:
            apogee_event = first_event
            if second_event is not None and second_event["distance"] < 390000:
                perigee_event = second_event
            elif second_event is not None:
                apogee_event = second_event
        else:
            perigee_event = first_event
            if second_event is not None and second_event["distance"] >= 390000:
                apogee_event = second_event
            elif second_event is not None:
                perigee_event = second_event

        rows.append({"year": current_year, "perigee": perigee_event, "apogee": apogee_event})

    return rows


def compute_perigee_apogee_minmax_flags(rows: List[dict]) -> None:
    perigee_by_year: Dict[int, List[dict]] = defaultdict(list)
    apogee_by_year: Dict[int, List[dict]] = defaultdict(list)

    for row in rows:
        year = row["year"]
        if row.get("perigee"):
            perigee_by_year[year].append(row["perigee"])
        if row.get("apogee"):
            apogee_by_year[year].append(row["apogee"])

    for events_by_year in (perigee_by_year, apogee_by_year):
        for year, events in events_by_year.items():
            if not events:
                continue
            dists = [e["distance"] for e in events]
            dmin = min(dists)
            dmax = max(dists)
            for e in events:
                if e["distance"] == dmax:
                    e["minmax"] = "M"
                elif e["distance"] == dmin:
                    e["minmax"] = "m"
                else:
                    e["minmax"] = ""


def write_perigees_csv(path: Path, rows: List[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Perigee", "", "", "", "", "Apogee", "", "", "", "", "", "", "", "", "", "", "", ""])
        writer.writerow(["Date", "UTC", "Distance", "MinMax", "", "Date", "UTC", "Distance", "MinMax", "", "", "", "", "", "", "", "", ""])

        for entry in rows:
            year = entry["year"]
            row = [""] * 18

            p = entry.get("perigee")
            if p:
                row[0] = date_label(year, p["month"], p["day"])
                row[1] = p["time"]
                row[2] = str(p["distance"])
                row[3] = p.get("minmax", "")

            a = entry.get("apogee")
            if a:
                row[5] = date_label(year, a["month"], a["day"])
                row[6] = a["time"]
                row[7] = str(a["distance"])
                row[8] = a.get("minmax", "")

            writer.writerow(row)


def write_moonphases_csv(
    path: Path,
    phases_by_year: Dict[int, Dict[str, List[dict]]],
    solar_map: Dict[Tuple[int, int, int], str],
    lunar_map: Dict[Tuple[int, int, int], str],
) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["New Moon", "", "", "", "First Quarter", "", "", "", "Full Moon", "", "", "", "Last Quarter", ""])
        writer.writerow(["Date", "UTC", "Eclipse", "", "Date", "UTC", "", "", "Date", "UTC", "Eclipse", "", "Date", "UTC"])

        for year in range(YEAR_START, YEAR_END + 1):
            y = phases_by_year.get(year, {"new": [], "first": [], "full": [], "last": []})
            n = y.get("new", [])
            q1 = y.get("first", [])
            full = y.get("full", [])
            q3 = y.get("last", [])

            row_count = max(len(n), len(q1), len(full), len(q3))
            for i in range(row_count):
                row = [""] * 14

                if i < len(n):
                    e = n[i]
                    row[0] = date_label(*e["date"])
                    row[1] = e["time"]
                    row[2] = normalize_phase_eclipse("new", e["date"], e["raw_code"], solar_map, lunar_map)

                if i < len(q1):
                    e = q1[i]
                    row[4] = date_label(*e["date"])
                    row[5] = e["time"]

                if i < len(full):
                    e = full[i]
                    row[8] = date_label(*e["date"])
                    row[9] = e["time"]
                    row[10] = normalize_phase_eclipse("full", e["date"], e["raw_code"], solar_map, lunar_map)

                if i < len(q3):
                    e = q3[i]
                    row[12] = date_label(*e["date"])
                    row[13] = e["time"]

                writer.writerow(row)


def calculate_planetary_events(ephemeris_path: Path) -> List[dict]:
    if load is None or almanac is None:
        raise AstroDataError("skyfield is required for elongation/opposition generation. Install it with: pip install skyfield")

    ts = load.timescale()
    eph = load(str(ephemeris_path))

    earth = eph["earth"]
    sun = eph["sun"]

    def to_ts(ts_local, d: dt.datetime):
        sec = d.second + d.microsecond / 1_000_000.0
        return ts_local.utc(d.year, d.month, d.day, d.hour, d.minute, sec)

    def elongation_deg(t, body) -> float:
        planet_vec = earth.at(t).observe(body).apparent().position.km
        sun_vec = earth.at(t).observe(sun).apparent().position.km
        dot = float(np.dot(planet_vec, sun_vec))
        denom = float(np.linalg.norm(planet_vec) * np.linalg.norm(sun_vec))
        if denom == 0.0:
            return 0.0
        cosang = max(-1.0, min(1.0, dot / denom))
        return math.degrees(math.acos(cosang))

    def east_west(t, body) -> str:
        _, sun_lon, _ = earth.at(t).observe(sun).apparent().ecliptic_latlon()
        _, planet_lon, _ = earth.at(t).observe(body).apparent().ecliptic_latlon()
        delta = (planet_lon.degrees - sun_lon.degrees + 540.0) % 360.0 - 180.0
        return "E" if delta > 0 else "W"

    events: List[dict] = []

    start = ts.utc(YEAR_START, 1, 1)
    stop = ts.utc(YEAR_END + 1, 1, 1)

    outer_planets = [
        ("Mars", "mars barycenter"),
        ("Jupiter", "jupiter barycenter"),
        ("Saturn", "saturn barycenter"),
        ("Uranus", "uranus barycenter"),
        ("Neptune", "neptune barycenter"),
    ]

    inner_planets = [
        ("Mercury", "mercury"),
        ("Venus", "venus"),
    ]

    total_planets = len(outer_planets) + len(inner_planets)
    planet_idx = 0
    for label, key in outer_planets:
        planet_idx += 1
        print(f"  [{planet_idx}/{total_planets}] Computing oppositions for {label} ...", end=" ", flush=True)
        body = eph[key]
        f = almanac.oppositions_conjunctions(eph, body)
        times, values = almanac.find_discrete(start, stop, f)
        count = 0
        for t, v in zip(times, values):
            if int(v) != 1:
                continue
            utc_dt = t.utc_datetime().replace(tzinfo=None)
            if utc_dt.year < YEAR_START or utc_dt.year > YEAR_END:
                continue
            events.append(
                {
                    "planet": label,
                    "event": "Opposition",
                    "direction": "",
                    "datetime": utc_dt,
                    "angle": elongation_deg(t, body),
                }
            )
            count += 1
        print(f"done ({count} events)")

    def refine_peak(ts_local, body, center_dt: dt.datetime) -> Tuple[dt.datetime, float]:
        step_days = 0.5
        t_prev = to_ts(ts_local, center_dt - dt.timedelta(days=step_days))
        t_mid = to_ts(ts_local, center_dt)
        t_next = to_ts(ts_local, center_dt + dt.timedelta(days=step_days))

        y_prev = elongation_deg(t_prev, body)
        y_mid = elongation_deg(t_mid, body)
        y_next = elongation_deg(t_next, body)

        denom = (y_prev - 2.0 * y_mid + y_next)
        if abs(denom) < 1e-12:
            return center_dt, y_mid

        offset_steps = 0.5 * (y_prev - y_next) / denom
        offset_steps = max(-1.0, min(1.0, offset_steps))
        refined_dt = center_dt + dt.timedelta(days=offset_steps * step_days)
        refined_t = to_ts(ts_local, refined_dt)
        return refined_dt, elongation_deg(refined_t, body)

    grid_start = dt.datetime(YEAR_START, 1, 1) - dt.timedelta(days=2)
    grid_stop = dt.datetime(YEAR_END + 1, 1, 1) + dt.timedelta(days=2)
    half_day = dt.timedelta(hours=12)

    sample_dts: List[dt.datetime] = []
    cursor = grid_start
    while cursor <= grid_stop:
        sample_dts.append(cursor)
        cursor += half_day

    years = [d.year for d in sample_dts]
    months = [d.month for d in sample_dts]
    days = [d.day for d in sample_dts]
    hours = [d.hour for d in sample_dts]
    minutes = [d.minute for d in sample_dts]
    seconds = [d.second + d.microsecond / 1_000_000.0 for d in sample_dts]
    t_grid = ts.utc(years, months, days, hours, minutes, seconds)

    for label, key in inner_planets:
        planet_idx += 1
        print(f"  [{planet_idx}/{total_planets}] Computing elongations for {label} ...", end=" ", flush=True)
        body = eph[key]
        count = 0

        planet_vec = earth.at(t_grid).observe(body).apparent().position.km
        sun_vec = earth.at(t_grid).observe(sun).apparent().position.km
        dot = np.sum(planet_vec * sun_vec, axis=0)
        denom = np.linalg.norm(planet_vec, axis=0) * np.linalg.norm(sun_vec, axis=0)
        cosang = np.clip(np.divide(dot, denom, out=np.zeros_like(dot), where=denom != 0), -1.0, 1.0)
        sample_vals = np.degrees(np.arccos(cosang))

        for i in range(1, len(sample_vals) - 1):
            if sample_vals[i] <= sample_vals[i - 1] or sample_vals[i] <= sample_vals[i + 1]:
                continue

            refined_dt, refined_angle = refine_peak(ts, body, sample_dts[i])
            if refined_dt.year < YEAR_START or refined_dt.year > YEAR_END:
                continue

            direction = east_west(to_ts(ts, refined_dt), body)
            events.append(
                {
                    "planet": label,
                    "event": "Greatest Elongation",
                    "direction": direction,
                    "datetime": refined_dt,
                    "angle": refined_angle,
                }
            )
            count += 1
        print(f"done ({count} events)")

    events.sort(key=lambda e: e["datetime"])
    return events


def calculate_meteor_peaks(ephemeris_path: Path) -> List[dict]:
    """Compute annual peak UTC times for major meteor showers (2001-2100).

    For each shower the Sun's ecliptic longitude is tracked using the JPL
    ephemeris and the peak is located by iterative Newton-style refinement
    around the catalogued solar longitude value.
    """
    if load is None:
        raise AstroDataError(
            "skyfield is required for meteor peak computation. "
            "Install it with: pip install skyfield"
        )

    ts_obj = load.timescale()
    eph = load(str(ephemeris_path))
    earth = eph["earth"]
    sun = eph["sun"]

    SOLAR_RATE = 360.0 / 365.25  # approximate degrees per day

    years = list(range(YEAR_START, YEAR_END + 1))
    n_years = len(years)
    events: List[dict] = []
    total = len(METEOR_SHOWERS)

    for shower_idx, (name, approx_month, approx_day, peak_lon, zhr) in enumerate(
        METEOR_SHOWERS, 1
    ):
        print(
            f"  [{shower_idx}/{total}] Computing peaks for {name} ...",
            end=" ",
            flush=True,
        )

        # Initial approximate dates for every year
        current_dts = [
            dt.datetime(y, approx_month, approx_day, 12, 0, 0) for y in years
        ]

        # Three refinement passes (converges to sub-minute accuracy)
        for _ in range(3):
            t_arr = ts_obj.utc(
                [d.year for d in current_dts],
                [d.month for d in current_dts],
                [d.day for d in current_dts],
                [d.hour for d in current_dts],
                [d.minute for d in current_dts],
                [d.second + d.microsecond / 1_000_000.0 for d in current_dts],
            )
            apparent = earth.at(t_arr).observe(sun).apparent()
            _, lon_arr, _ = apparent.ecliptic_latlon()
            lon_deg = lon_arr.degrees

            # Signed angular difference (handles 360/0 wrap)
            diff = (lon_deg - peak_lon + 180.0) % 360.0 - 180.0
            shift_days = -diff / SOLAR_RATE

            current_dts = [
                d + dt.timedelta(days=float(s))
                for d, s in zip(current_dts, shift_days)
            ]

        for y, d in zip(years, current_dts):
            events.append(
                {
                    "shower": name,
                    "year": y,
                    "datetime": d,
                    "zhr": zhr,
                }
            )

        print(f"done ({n_years} peaks)")

    events.sort(key=lambda e: e["datetime"])
    return events


def write_meteors_csv(path: Path, meteor_events: List[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Shower", "Date", "UTC", "ZHR"])

        for e in meteor_events:
            d = e["datetime"]
            writer.writerow([
                e["shower"],
                date_label(d.year, d.month, d.day),
                f"{d:%H:%M}",
                e["zhr"],
            ])


def write_elongation_csv(path: Path, events: List[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Planet", "Event", "Direction", "Date", "UTC", "AngleDeg"])

        for e in events:
            ts_value = e["datetime"]
            writer.writerow(
                [
                    e["planet"],
                    e["event"],
                    e["direction"],
                    date_label(ts_value.year, ts_value.month, ts_value.day),
                    f"{ts_value:%H:%M}",
                    f"{e['angle']:.2f}",
                ]
            )


def write_combined_xlsx(xlsx_path: Path, csv_paths: List[Tuple[str, Path]]) -> None:
    if Workbook is None:
        raise AstroDataError(
            "openpyxl is required for --xlsx output. Install it with: pip install openpyxl"
        )

    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    for sheet_name, csv_path in csv_paths:
        ws = wb.create_sheet(title=sheet_name)
        with csv_path.open("r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)

        # Auto-fit column widths
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                val = row[0]
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    wb.save(str(xlsx_path))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate 21st-century perigees.csv and moonphases.csv from live astronomy web sources."
    )
    parser.add_argument(
        "--out-dir",
        default=str(Path(__file__).resolve().parent),
        help="Directory where perigees.csv and moonphases.csv are written (default: script directory).",
    )
    parser.add_argument(
        "--download-dir",
        default=str(Path(__file__).resolve().parent / "downloads"),
        help="Directory where downloaded source HTML is saved.",
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help="Use only saved HTML in --download-dir and skip web downloads.",
    )
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Also combine all CSVs into a single astro.xlsx workbook.",
    )
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    download_dir = Path(args.download_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    download_dir.mkdir(parents=True, exist_ok=True)

    t_total = time.time()
    step = 0
    total_steps = 9 if args.xlsx else 8

    def progress(msg: str) -> None:
        nonlocal step
        step += 1
        elapsed = time.time() - t_total
        print(f"[{step}/{total_steps}] {msg}  (elapsed {elapsed:.1f}s)")

    # --- Step 1: Load sources ---
    if args.offline:
        progress("Loading cached HTML sources ...")
        sources = load_sources_offline(download_dir)
    else:
        progress("Downloading HTML sources ...")
        sources = fetch_and_save_sources(download_dir)

    # --- Step 2: Parse eclipse catalogs ---
    progress("Parsing NASA eclipse catalogs ...")
    solar_map = parse_nasa_solar_types(sources["nasa_solar_eclipses_2001_2100.html"])
    lunar_map = parse_nasa_lunar_types(sources["nasa_lunar_eclipses_2001_2100.html"])
    print(f"  Solar eclipses: {len(solar_map)}, Lunar eclipses: {len(lunar_map)}")

    if not solar_map or not lunar_map:
        raise AstroDataError("Failed to parse NASA eclipse catalogs.")

    # --- Step 3: Parse moon phases ---
    progress("Parsing moon phases ...")
    phases_by_year = parse_phase_lines(sources["astropixels_phases2001gmt.html"])
    if not phases_by_year:
        raise AstroDataError("Failed to parse moon phase source table.")
    print(f"  Years with phase data: {len(phases_by_year)}")

    # --- Step 4: Parse perigee/apogee ---
    progress("Parsing perigee/apogee data ...")
    perigee_apogee_rows = parse_perigee_apogee_rows(sources["astropixels_moonperap2001.html"])
    if not perigee_apogee_rows:
        raise AstroDataError("Failed to parse perigee/apogee source table.")
    print(f"  Perigee/apogee rows: {len(perigee_apogee_rows)}")

    # --- Step 5: Compute planetary events (slowest step) ---
    progress("Computing planetary elongations & oppositions ...")
    ephemeris_path = ensure_jpl_ephemeris(download_dir, args.offline)
    elongation_events = calculate_planetary_events(ephemeris_path)
    if not elongation_events:
        raise AstroDataError("Failed to generate planetary elongation/opposition events.")
    print(f"  Planetary events: {len(elongation_events)}")

    # --- Step 6: Compute meteor peaks ---
    progress("Computing meteor shower peaks ...")
    meteor_events = calculate_meteor_peaks(ephemeris_path)
    if not meteor_events:
        raise AstroDataError("Failed to compute meteor shower peaks.")
    print(f"  Meteor shower peaks: {len(meteor_events)}")

    # --- Step 7: Compute min/max flags ---
    progress("Computing perigee/apogee min/max flags ...")
    compute_perigee_apogee_minmax_flags(perigee_apogee_rows)

    # --- Step 8: Write CSVs ---
    progress("Writing CSV files ...")
    write_perigees_csv(out_dir / "perigees.csv", perigee_apogee_rows)
    write_moonphases_csv(out_dir / "moonphases.csv", phases_by_year, solar_map, lunar_map)
    write_elongation_csv(out_dir / "elongation.csv", elongation_events)
    write_meteors_csv(out_dir / "meteors.csv", meteor_events)

    if args.xlsx:
        progress("Writing astro.xlsx ...")
        xlsx_path = out_dir / "astro.xlsx"
        write_combined_xlsx(
            xlsx_path,
            [
                ("Moon Phases", out_dir / "moonphases.csv"),
                ("Perigees & Apogees", out_dir / "perigees.csv"),
                ("Elongation", out_dir / "elongation.csv"),
                ("Meteors", out_dir / "meteors.csv"),
            ],
        )

    elapsed_total = time.time() - t_total
    print(f"\nDone in {elapsed_total:.1f}s.")
    if args.offline:
        print(f"Loaded downloads from: {download_dir}")
    else:
        print(f"Saved downloads to: {download_dir}")
    print(f"Wrote: {out_dir / 'perigees.csv'}")
    print(f"Wrote: {out_dir / 'moonphases.csv'}")
    print(f"Wrote: {out_dir / 'elongation.csv'}")
    print(f"Wrote: {out_dir / 'meteors.csv'}")
    if args.xlsx:
        print(f"Wrote: {xlsx_path}")


if __name__ == "__main__":
    main()
