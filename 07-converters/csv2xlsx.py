#!/usr/bin/env python3
"""
csv2xlsx.py

Stream-convert a large CSV (e.g., 500MB+) into XLSX.
Optionally select N random rows (uniform) via reservoir sampling.

Features:
- Proper CSV parsing with quoted fields using Python's csv module
- Memory-friendly XLSX writing using openpyxl write_only mode
- Optional sampling: --sample N
- Optional splitting into multiple sheets if row count exceeds Excel limits
- Optional chunking into multiple XLSX files: --chunk-rows N

Usage:
1) Convert the entire CSV → XLSX (streaming)
python csv2xlsx.py big.csv big.xlsx

2) Sample 10,000 random rows (uniform) → XLSX
python csv2xlsx.py big.csv sample.xlsx --sample 10000 --seed 42

3) If the CSV might exceed Excel’s row limit, split into sheets
python csv2xlsx.py big.csv big.xlsx --split-sheets

4) Write multiple numbered XLSX files (chunks) into <input>_chunks folder
python csv2xlsx.py big.csv ignored.xlsx --chunk-rows 200000
"""

from __future__ import annotations

import argparse
import csv
import io
import os
import random
import sys
import time
from typing import Any, List, Optional, Sequence, Tuple

from openpyxl import Workbook

EXCEL_MAX_ROWS = 1_048_576  # includes header row if you write one
PROGRESS_EVERY_ROWS = 100_000
TICK_EVERY_ROWS = 100


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Convert large CSV to XLSX (streaming). Optionally sample random rows."
    )
    p.add_argument("input_csv", help="Path to input CSV file")
    p.add_argument("output_xlsx", help="Path to output XLSX file")

    p.add_argument("--delimiter", default=",", help="CSV delimiter (default: ,)")
    p.add_argument("--quotechar", default='"', help='CSV quote character (default: ")')
    p.add_argument("--encoding", default="utf-8", help="File encoding (default: utf-8)")
    p.add_argument(
        "--errors",
        default="replace",
        help="Encoding error handling (default: replace). Try 'strict' if you want failures.",
    )

    p.add_argument(
        "--no-header",
        action="store_true",
        help="Treat first row as data (do not treat as header). Default assumes header exists.",
    )

    p.add_argument(
        "--sample",
        type=int,
        default=0,
        help="If >0, randomly sample N rows (uniform) from the CSV (excluding header by default).",
    )
    p.add_argument(
        "--seed",
        type=int,
        default=None,
        help="Random seed for sampling (default: None). Use for reproducible samples.",
    )

    p.add_argument(
        "--sheet-name",
        default="Sheet1",
        help="Base sheet name (default: Sheet1). If split, names become Sheet1_1, Sheet1_2, ...",
    )
    p.add_argument(
        "--split-sheets",
        action="store_true",
        help="Split output across multiple sheets if rows exceed Excel limit.",
    )

    p.add_argument(
        "--max-rows-per-sheet",
        type=int,
        default=EXCEL_MAX_ROWS,
        help=f"Max rows per sheet (default: {EXCEL_MAX_ROWS}). Must be <= {EXCEL_MAX_ROWS}.",
    )

    p.add_argument(
        "--chunk-rows",
        type=int,
        default=0,
        help=(
            "If >0, write output as multiple numbered XLSX files with this many data rows per file. "
            "Files are written to a folder named <input_stem>_chunks next to the input CSV. "
            "Cannot be combined with --split-sheets."
        ),
    )

    return p.parse_args()


def open_csv_reader(
    path: str,
    delimiter: str,
    quotechar: str,
    encoding: str,
    errors: str,
) -> Tuple[Any, io.TextIOWrapper]:

    f = open(path, "r", encoding=encoding, errors=errors, newline="")
    reader = csv.reader(
        f,
        delimiter=delimiter,
        quotechar=quotechar,
        doublequote=True,
        skipinitialspace=False,
    )
    return reader, f


def reservoir_sample_rows(
    reader: Any,
    sample_n: int,
    header: Optional[List[str]],
    rng: random.Random,
) -> List[List[str]]:
    """
    Uniformly sample sample_n rows from the stream using reservoir sampling.
    """
    reservoir: List[List[str]] = []
    seen = 0

    for row in reader:
        seen += 1
        if len(reservoir) < sample_n:
            reservoir.append(row)
        else:
            j = rng.randrange(seen)  # [0, seen-1]
            if j < sample_n:
                reservoir[j] = row

    return reservoir


def make_sheet_name(base: str, idx: int, split: bool) -> str:
    return f"{base}_{idx}" if split else base


def write_xlsx_streaming(
    input_csv: str,
    output_xlsx: str,
    delimiter: str,
    quotechar: str,
    encoding: str,
    errors: str,
    has_header: bool,
    sample_n: int,
    seed: Optional[int],
    sheet_base: str,
    split_sheets: bool,
    max_rows_per_sheet: int,
    chunk_rows: int = 0,
) -> None:
    if max_rows_per_sheet > EXCEL_MAX_ROWS:
        raise ValueError(f"--max-rows-per-sheet must be <= {EXCEL_MAX_ROWS}")

    if chunk_rows and chunk_rows > 0 and split_sheets:
        raise ValueError("--chunk-rows cannot be combined with --split-sheets")

    started_at = time.monotonic()
    started_wall = time.strftime("%Y-%m-%d %H:%M:%S")

    try:
        input_size = os.path.getsize(input_csv)
    except OSError:
        input_size = None

    print(f"csv2xlsx: START {started_wall}")
    if input_size is not None:
        print(f"csv2xlsx: input={input_csv} ({input_size:,} bytes)")
    else:
        print(f"csv2xlsx: input={input_csv}")
    print(f"csv2xlsx: output={output_xlsx}")
    print(
        "csv2xlsx: mode="
        + (f"sample({sample_n})" if sample_n and sample_n > 0 else "full")
        + (" split-sheets" if split_sheets else "")
        + (f" chunk-rows({chunk_rows})" if chunk_rows and chunk_rows > 0 else "")
    )
    print(
        "csv2xlsx: csv settings="
        + f"delimiter={repr(delimiter)} quotechar={repr(quotechar)} encoding={encoding} errors={errors}"
    )

    rng = random.Random(seed)

    reader, f = open_csv_reader(input_csv, delimiter, quotechar, encoding, errors)
    try:
        header: Optional[List[str]] = None
        if has_header:
            try:
                header = next(reader)
            except StopIteration:
                raise ValueError("CSV appears to be empty.")

        if header is not None:
            print(f"csv2xlsx: header columns={len(header)}")
        else:
            print("csv2xlsx: no header")

        if chunk_rows and chunk_rows > 0:
            max_data_rows = EXCEL_MAX_ROWS - (1 if header is not None else 0)
            if chunk_rows > max_data_rows:
                raise ValueError(
                    f"--chunk-rows must be <= {max_data_rows} (so the file stays within Excel's row limit)"
                )
            print("csv2xlsx: chunk mode enabled; output_xlsx argument is ignored")

        chunks_dir: Optional[str] = None
        chunk_idx = 0
        wb: Optional[Workbook] = None
        ws: Any = None

        def chunks_output_dir() -> str:
            nonlocal chunks_dir
            if chunks_dir is None:
                input_dir = os.path.dirname(os.path.abspath(input_csv))
                stem = os.path.splitext(os.path.basename(input_csv))[0]
                chunks_dir = os.path.join(input_dir, f"{stem}_chunks")
            return chunks_dir

        def chunk_path(idx: int) -> str:
            out_dir = chunks_output_dir()
            stem = os.path.splitext(os.path.basename(input_csv))[0]
            return os.path.join(out_dir, f"{stem}_{idx:04d}.xlsx")

        sheet_idx = 1

        def start_new_workbook():
            nonlocal wb, ws, chunk_idx, sheet_idx
            if chunk_rows and chunk_rows > 0:
                chunk_idx += 1
                out_dir = chunks_output_dir()
                if not os.path.exists(out_dir):
                    os.makedirs(out_dir, exist_ok=True)
                    print(f"csv2xlsx: created chunks folder {out_dir}")
                wb = Workbook(write_only=True)
                sheet_idx = 1
                ws = wb.create_sheet(title=sheet_base)
                print(f"csv2xlsx: writing chunk {chunk_idx} file={chunk_path(chunk_idx)}")
            else:
                if wb is None:
                    wb = Workbook(write_only=True)
                ws = wb.create_sheet(title=make_sheet_name(sheet_base, sheet_idx, split_sheets))
                print(f"csv2xlsx: writing sheet {sheet_idx} name={ws.title}")

        start_new_workbook()

        # row_counter counts rows written to current sheet (including header if written)
        row_counter = 0
        total_rows_written = 0
        total_rows_read = 0
        last_pulse_at = started_at
        last_pulse_rows = 0
        dot_active = False

        def tick():
            nonlocal dot_active
            print(".", end="", flush=True)
            dot_active = True

        def tick_newline():
            nonlocal dot_active
            if dot_active:
                print("")
                dot_active = False

        def pulse(tag: str):
            nonlocal last_pulse_at, last_pulse_rows
            tick_newline()
            now = time.monotonic()
            elapsed = now - started_at
            delta_rows = total_rows_read - last_pulse_rows
            delta_t = max(now - last_pulse_at, 1e-9)
            rate = delta_rows / delta_t
            extra = ""
            if chunk_rows and chunk_rows > 0:
                extra = f" chunk={chunk_idx}"
            print(
                f"csv2xlsx: {tag}: read={total_rows_read:,} written={total_rows_written:,}"
                f"{extra} sheet={sheet_idx} sheet_rows={row_counter:,} elapsed={elapsed:,.1f}s rate={rate:,.1f} rows/s"
            )
            last_pulse_at = now
            last_pulse_rows = total_rows_read

        def maybe_write_header():
            nonlocal row_counter, total_rows_written
            if header is not None:
                ws.append(header)
                row_counter += 1
                total_rows_written += 1

        maybe_write_header()

        if sample_n and sample_n > 0:
            # Sample rows (excluding header by default because header already consumed)
            print(f"csv2xlsx: sampling {sample_n:,} rows (reservoir)")
            sampled = reservoir_sample_rows(reader, sample_n, header, rng)

            total_rows_read = len(sampled)
            pulse("sampled")

            for row in sampled:
                # Chunk or split handling
                if chunk_rows and chunk_rows > 0:
                    # chunk_rows counts data rows (excluding header)
                    if (row_counter - (1 if header is not None else 0)) >= chunk_rows:
                        tick_newline()
                        assert wb is not None
                        out_path = chunk_path(chunk_idx)
                        print(f"csv2xlsx: saving chunk {chunk_idx} -> {out_path}")
                        wb.save(out_path)
                        start_new_workbook()
                        row_counter = 0
                        maybe_write_header()
                elif split_sheets and row_counter >= max_rows_per_sheet:
                    tick_newline()
                    sheet_idx += 1
                    start_new_workbook()
                    row_counter = 0
                    maybe_write_header()

                ws.append(row)
                row_counter += 1
                total_rows_written += 1

                if total_rows_written % TICK_EVERY_ROWS == 0:
                    tick()

                if total_rows_written % PROGRESS_EVERY_ROWS == 0:
                    pulse("progress")

        else:
            # Full streaming conversion
            for row in reader:
                total_rows_read += 1
                if total_rows_read % TICK_EVERY_ROWS == 0:
                    tick()

                # Chunk or split handling
                if chunk_rows and chunk_rows > 0:
                    if (row_counter - (1 if header is not None else 0)) >= chunk_rows:
                        tick_newline()
                        assert wb is not None
                        out_path = chunk_path(chunk_idx)
                        print(f"csv2xlsx: saving chunk {chunk_idx} -> {out_path}")
                        wb.save(out_path)
                        start_new_workbook()
                        row_counter = 0
                        maybe_write_header()
                elif split_sheets and row_counter >= max_rows_per_sheet:
                    tick_newline()
                    sheet_idx += 1
                    start_new_workbook()
                    row_counter = 0
                    maybe_write_header()

                ws.append(row)
                row_counter += 1
                total_rows_written += 1

                if total_rows_read % PROGRESS_EVERY_ROWS == 0:
                    pulse("progress")

            pulse("read-complete")

            tick_newline()

        assert wb is not None
        tick_newline()

        # Save output
        if chunk_rows and chunk_rows > 0:
            out_path = chunk_path(chunk_idx)
            print(f"csv2xlsx: saving final chunk {chunk_idx} -> {out_path}")
            save_started = time.monotonic()
            wb.save(out_path)
            save_elapsed = time.monotonic() - save_started
        else:
            # If user didn't enable split and we exceed Excel limits, warn (Excel may not open it properly)
            if not split_sheets:
                total_rows_possible = row_counter
                if total_rows_possible > EXCEL_MAX_ROWS:
                    print(
                        f"WARNING: Wrote {total_rows_possible} rows to one sheet, "
                        f"which exceeds Excel's max rows ({EXCEL_MAX_ROWS}). "
                        f"Re-run with --split-sheets.",
                        file=sys.stderr,
                    )

            # Ensure output folder exists
            out_dir = os.path.dirname(os.path.abspath(output_xlsx))
            if out_dir and not os.path.exists(out_dir):
                os.makedirs(out_dir, exist_ok=True)
                print(f"csv2xlsx: created output folder {out_dir}")

            print("csv2xlsx: saving workbook...")
            save_started = time.monotonic()
            wb.save(output_xlsx)
            save_elapsed = time.monotonic() - save_started

        total_elapsed = time.monotonic() - started_at
        print(
            f"csv2xlsx: DONE rows_written={total_rows_written:,} sheets={sheet_idx} "
            f"save_time={save_elapsed:,.1f}s total_time={total_elapsed:,.1f}s"
        )

    finally:
        f.close()


def main() -> int:
    args = parse_args()

    try:
        write_xlsx_streaming(
            input_csv=args.input_csv,
            output_xlsx=args.output_xlsx,
            delimiter=args.delimiter,
            quotechar=args.quotechar,
            encoding=args.encoding,
            errors=args.errors,
            has_header=(not args.no_header),
            sample_n=args.sample,
            seed=args.seed,
            sheet_base=args.sheet_name,
            split_sheets=args.split_sheets,
            max_rows_per_sheet=args.max_rows_per_sheet,
            chunk_rows=args.chunk_rows,
        )
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
